import pandas as pd
import os
import sys
import time
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# TODO: load chunks for better performance

# Loads the Excel file
excel_file = sys.argv[1]
new_file_path = r"./created_sheets"

# Headers to search against
insertHeader = sys.argv[2]
dataHeader = sys.argv[3]

# Directory with images
img_folder = sys.argv[4]

# Handled file extensions for images
extensions = ['png', 'jpg', 'jpeg']

insertedImages = 0
missingImages = 0

start_time = time.time()

shutil.copy(excel_file, new_file_path)

try:
    workbook = load_workbook(new_file_path)
except Exception as e:
    print(f"Error loading the Excel file: {e}")
    sys.exit(1)

for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]
    print(f"Processing sheet: {sheet_name}")

    data = pd.DataFrame(ws.values)

    # Header starting at 0
    header = data.iloc[0]
    data = data[1:]
    data.columns = header

    # Find columns based on name
    try:
        data_col = header[header.str.contains(
            dataHeader, case=False, na=False)].index[0]
        insert_image_col = header[header.str.contains(
            insertHeader, case=False, na=False)].index[0]
    except IndexError:
        print(f"Error: Couldn't find columns for '{dataHeader}' or '{
              insertHeader}' in sheet: {sheet_name}")
        continue

    for index, row in data.iterrows():
        value = row[data_col]
        # TODO: handle other extensions
        img_path = os.path.join(img_folder, f"{value}.png")

        if value is None:
            continue

        if os.path.exists(img_path):
            img = Image(img_path)
            img.width = 75
            img.height = 75
            cell = ws.cell(row=index + 1, column=insert_image_col + 1)
            img.anchor = cell.coordinate
            ws.add_image(img)
            insertedImages += 1
            print(f"Inserted image for {dataHeader}: {value}")
        else:
            print(f"Image not found for {dataHeader}: {value}")
            cell = ws.cell(row=index + 1, column=insert_image_col + 1)
            cell.value = "Image not found"

            missingImages += 1

try:
    workbook.save(new_file_path)
except Exception as e:
    print(f"Error saving the Excel file: {e}")

elapsed_time = time.time() - start_time

print(f"All sheets processed and images inserted with {
      insertedImages} inserted and {missingImages} missing. Execution time: {elapsed_time:.2f} seconds.")
